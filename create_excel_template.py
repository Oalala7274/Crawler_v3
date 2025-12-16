"""
创建 Excel 控制台模板文件
运行此脚本前请先安装依赖: pip install openpyxl
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sources'

    # 表头
    headers = ['Base_URL', 'Is_Enabled', 'Parser_Name', 'Action_Name', 'Search_Name', 'Filter_Name', 'Engine']
    
    # 表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 示例数据
    sample_data = [
        ['https://news.google.com/search?q=Separator&hl=en', 'TRUE', 'google_news', 'WAIT', 'NONE', '', 'edge'],
        ['https://www.reuters.com/site-search/?query=Separator', 'TRUE', 'REUTERS_SEARCH', 'INFINITE_SCROLL', 'SEARCH_Q', '', 'edge'],
        ['https://www.argusmedia.com/en/search?q=Separator', 'FALSE', 'ARGUS_MEDIA_SEARCH', 'WAIT', 'SEARCH_Q', '', 'edge'],
        ['https://www.bing.com/news/search?q=Separator', 'TRUE', 'BING_NEWS_SEARCH', 'WAIT', 'SEARCH_Q', '', 'edge'],
        ['https://www.metal.com/news?keyword=Separator', 'TRUE', 'METAL_COM_SEARCH', 'WAIT', 'SEARCH_Q', '', 'edge'],
        ['https://batteriesnews.com/?s=Separator', 'TRUE', 'BATTERIES_NEWS_SEARCH', 'WAIT', 'SEARCH_Q', '', 'edge'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 调整列宽
    ws.column_dimensions['A'].width = 65
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10

    # 保存文件
    filename = '国际隔膜信息源.xlsx'
    wb.save(filename)
    print(f'✓ Excel模板已创建: {filename}')
    print(f'  包含 {len(sample_data)} 条示例配置')
    print('\n请根据需要修改配置后运行 main.py')

if __name__ == '__main__':
    create_template()

