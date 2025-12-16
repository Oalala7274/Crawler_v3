"""
A模块：准备任务
读取Excel控制台，生成JSONL格式的任务列表
"""
import os
import uuid
from typing import List
import openpyxl

from utils.models import CrawlTask
from utils.config_loader import ConfigLoader

# 临时文件目录
TEMP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'temp')


def run(config: dict) -> List[CrawlTask]:
    """
    读取Excel并生成任务列表
    
    Args:
        config: 配置字典
        
    Returns:
        CrawlTask列表
        
    同时写入 temp/tasks.jsonl
    """
    # 确保临时目录存在
    os.makedirs(TEMP_DIR, exist_ok=True)
    
    # 获取Excel文件路径
    excel_file = config['settings'].get('SOURCE_EXCEL_FILE', '国际隔膜信息源.xlsx')
    
    # 尝试在多个位置查找Excel文件
    possible_paths = [
        excel_file,
        os.path.join(os.path.dirname(os.path.dirname(__file__)), excel_file),
        os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', excel_file),
    ]
    
    excel_path = None
    for path in possible_paths:
        if os.path.exists(path):
            excel_path = path
            break
    
    if not excel_path:
        print(f"❌ 找不到Excel文件: {excel_file}")
        print(f"   请确保文件存在于项目根目录")
        return []
    
    print(f"📄 读取Excel: {excel_path}")
    
    # 读取Excel
    tasks = []
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active
        
        # 获取表头
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # 查找列索引
        col_map = {}
        required_cols = ['Base_URL', 'Is_Enabled', 'Parser_Name', 'Action_Name', 'Search_Name', 'Filter_Name', 'Engine']
        
        for i, header in enumerate(headers):
            for col in required_cols:
                if header.lower() == col.lower():
                    col_map[col] = i
                    break
        
        # 检查必需列
        missing_cols = [col for col in required_cols if col not in col_map]
        if missing_cols:
            print(f"⚠ Excel缺少列: {', '.join(missing_cols)}")
            # 尝试按位置猜测
            if len(headers) >= 7:
                col_map = {
                    'Base_URL': 0,
                    'Is_Enabled': 1,
                    'Parser_Name': 2,
                    'Action_Name': 3,
                    'Search_Name': 4,
                    'Filter_Name': 5,
                    'Engine': 6
                }
                print("   尝试按默认列顺序解析...")
        
        # 遍历数据行
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_map.get('Base_URL', 0)]:
                continue
            
            # 获取各列值
            base_url = str(row[col_map.get('Base_URL', 0)] or "").strip()
            is_enabled = str(row[col_map.get('Is_Enabled', 1)] or "").upper().strip()
            parser_name = str(row[col_map.get('Parser_Name', 2)] or "").strip()
            action_name = str(row[col_map.get('Action_Name', 3)] or "WAIT").strip()
            search_name = str(row[col_map.get('Search_Name', 4)] or "NONE").strip()
            filter_name = str(row[col_map.get('Filter_Name', 5)] or "").strip()
            engine = str(row[col_map.get('Engine', 6)] or "edge").strip()
            
            # 跳过未启用的行
            if is_enabled != "TRUE":
                continue
            
            row_count += 1
            
            # 获取关键词列表
            keywords = ConfigLoader.get_keyword_set(search_name, config)
            
            if keywords:
                # 需要替换关键词的URL
                for keyword in keywords:
                    # 替换URL中的占位符
                    url = base_url.replace("Separator", keyword)
                    url = url.replace("{keyword}", keyword)
                    url = url.replace("%s", keyword)
                    
                    # 对关键词进行URL编码
                    import urllib.parse
                    encoded_keyword = urllib.parse.quote(keyword)
                    url = url.replace("{keyword_encoded}", encoded_keyword)
                    
                    task = CrawlTask(
                        task_id=uuid.uuid4().hex[:8],
                        url=url,
                        parser_name=parser_name,
                        action_name=action_name,
                        filter_name=filter_name,
                        engine=engine
                    )
                    tasks.append(task)
            else:
                # 无需替换关键词，直接使用URL
                task = CrawlTask(
                    task_id=uuid.uuid4().hex[:8],
                    url=base_url,
                    parser_name=parser_name,
                    action_name=action_name,
                    filter_name=filter_name,
                    engine=engine
                )
                tasks.append(task)
        
        wb.close()
        print(f"   读取到 {row_count} 条有效配置")
        
    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        return []
    
    # 写入任务文件
    tasks_file = os.path.join(TEMP_DIR, 'tasks.jsonl')
    with open(tasks_file, 'w', encoding='utf-8') as f:
        for task in tasks:
            f.write(task.to_json() + '\n')
    
    print(f"   生成 {len(tasks)} 个爬取任务")
    print(f"   任务文件: {tasks_file}")
    
    return tasks


def load_tasks_from_file() -> List[CrawlTask]:
    """
    从临时文件加载任务列表
    
    Returns:
        CrawlTask列表
    """
    tasks_file = os.path.join(TEMP_DIR, 'tasks.jsonl')
    tasks = []
    
    if os.path.exists(tasks_file):
        with open(tasks_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    tasks.append(CrawlTask.from_json(line))
    
    return tasks

